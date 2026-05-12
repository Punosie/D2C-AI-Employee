"""
Run this once to generate sample_data.xlsx.
Upload to Google Drive and open as Google Sheets.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = Workbook()

HEADER_FILL = PatternFill("solid", fgColor="2D6A4F")
HEADER_FONT = Font(color="FFFFFF", bold=True)

def write_sheet(ws, headers, rows):
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append(row)
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 24


# ── Sheet 1: Inventory ────────────────────────────────────────────────────────

ws1 = wb.active
ws1.title = "Inventory"

write_sheet(ws1,
    headers=["sku", "product_name", "current_stock", "reorder_level", "reorder_qty", "unit_cost", "location", "last_updated"],
    rows=[
        ["VC-200", "Vanilla Soy Candle 200g",    42,  20, 100, 180, "Warehouse A", "2026-05-01"],
        ["LC-200", "Lavender Soy Candle 200g",   18,  20, 100, 180, "Warehouse A", "2026-05-01"],
        ["RC-200", "Rose Soy Candle 200g",        8,  20, 100, 185, "Warehouse A", "2026-05-01"],
        ["SC-200", "Sandalwood Candle 200g",     35,  20, 100, 190, "Warehouse A", "2026-05-01"],
        ["CC-200", "Citrus Burst Candle 200g",   55,  20, 100, 175, "Warehouse B", "2026-05-01"],
        ["GS-3",   "Gift Set (3 candles)",        12,  10,  50, 520, "Warehouse B", "2026-05-01"],
        ["GS-5",   "Gift Set (5 candles)",         6,  10,  30, 850, "Warehouse B", "2026-05-01"],
        ["VC-100", "Vanilla Soy Candle 100g",    70,  30, 150,  95, "Warehouse A", "2026-05-01"],
        ["LC-100", "Lavender Soy Candle 100g",   45,  30, 150,  95, "Warehouse A", "2026-05-01"],
        ["WC-200", "Woody Cedar Candle 200g",    22,  20, 100, 195, "Warehouse B", "2026-05-01"],
    ]
)


# ── Sheet 2: Raw Material Cost ────────────────────────────────────────────────

ws2 = wb.create_sheet("Raw Material Cost")

write_sheet(ws2,
    headers=["material", "vendor_name", "unit", "cost_per_unit", "monthly_usage", "monthly_cost", "last_purchase_date", "notes"],
    rows=[
        ["Soy Wax",              "GreenWax Supplies",    "kg",     220,  80,  17600, "2026-04-28", "Premium soy, low smoke"],
        ["Coconut Wax",          "NatureCraft Co.",      "kg",     310,  20,   6200, "2026-04-20", "Used in premium range"],
        ["Vanilla Fragrance Oil","Scent Studio India",   "100ml",  180, 120,  21600, "2026-04-25", "Top selling scent"],
        ["Lavender Essential Oil","PureAroma Exports",   "100ml",  240,  80,  19200, "2026-04-22", "Certified organic"],
        ["Rose Fragrance Oil",   "Scent Studio India",   "100ml",  210,  60,  12600, "2026-04-25", ""],
        ["Sandalwood Fragrance", "AromaBase India",      "100ml",  260,  55,  14300, "2026-04-18", "High retention scent"],
        ["Cotton Wicks",         "WickMart",             "pack of 100", 350, 15,  5250, "2026-04-30", "Pre-tabbed, 3in"],
        ["Glass Jars 200ml",     "PackRight Packaging",  "piece",   28, 500,  14000, "2026-04-15", "Clear glass, matte lid"],
        ["Glass Jars 100ml",     "PackRight Packaging",  "piece",   18, 300,   5400, "2026-04-15", ""],
        ["Kraft Boxes",          "EcoPack India",        "piece",   22, 600,  13200, "2026-04-10", "FSC certified"],
        ["Tissue Paper",         "EcoPack India",        "sheet",    2, 800,   1600, "2026-04-10", "Recycled"],
        ["Labels (custom print)","PrintBazaar",          "piece",    4, 600,   2400, "2026-04-12", "Waterproof matte"],
    ]
)


# ── Sheet 3: Vendors ──────────────────────────────────────────────────────────

ws3 = wb.create_sheet("Vendors")

write_sheet(ws3,
    headers=["vendor_name", "category", "contact_person", "phone", "email", "payment_terms", "lead_time_days", "rating", "notes"],
    rows=[
        ["GreenWax Supplies",     "Raw Material", "Arjun Mehta",    "9876500001", "arjun@greenwax.in",      "Net 30",   7,  5, "Primary soy wax supplier"],
        ["NatureCraft Co.",       "Raw Material", "Sneha Rao",      "9876500002", "sneha@naturecraft.in",   "Net 15",   10, 4, "Coconut wax, MOQ 10kg"],
        ["Scent Studio India",    "Raw Material", "Rohit Verma",    "9876500003", "rohit@scentstudio.in",   "Advance",  5,  5, "Best vanilla & rose oils"],
        ["PureAroma Exports",     "Raw Material", "Divya Nair",     "9876500004", "divya@purearoma.in",     "Net 30",   14, 4, "Organic certified"],
        ["AromaBase India",       "Raw Material", "Kiran Sharma",   "9876500005", "kiran@aromabase.in",     "Net 15",   7,  4, "Sandalwood specialist"],
        ["WickMart",              "Raw Material", "Priya Iyer",     "9876500006", "priya@wickmart.in",      "Advance",  3,  5, "Fast delivery"],
        ["PackRight Packaging",   "Packaging",    "Amit Joshi",     "9876500007", "amit@packright.in",      "Net 30",   10, 5, "Glass jars, bulk discount >500"],
        ["EcoPack India",         "Packaging",    "Ananya Gupta",   "9876500008", "ananya@ecopack.in",      "Net 15",   7,  4, "Sustainable packaging"],
        ["PrintBazaar",           "Packaging",    "Suresh Kumar",   "9876500009", "suresh@printbazaar.in",  "Advance",  5,  3, "Custom labels, min 500pcs"],
        ["Delhivery",             "Logistics",    "Rajesh Singh",   "9876500010", "rajesh@delhivery.com",   "Monthly",  1,  4, "Primary courier partner"],
        ["Shiprocket",            "Logistics",    "Account Manager","1800123456", "support@shiprocket.in",  "Monthly",  1,  4, "Aggregator, backup courier"],
    ]
)


# ── Sheet 4: Monthly Budget ───────────────────────────────────────────────────

ws4 = wb.create_sheet("Monthly Budget")

write_sheet(ws4,
    headers=["month", "category", "budgeted", "actual", "variance", "notes"],
    rows=[
        ["2026-04", "Raw Materials",      85000,  82400,   2600, "Under budget — bulk soy wax deal"],
        ["2026-04", "Packaging",          35000,  36800,  -1800, "Label reprint due to design change"],
        ["2026-04", "Meta Ads",           60000,  58200,   1800, ""],
        ["2026-04", "Google Ads",         40000,  43500,  -3500, "Increased bids for Mother's Day"],
        ["2026-04", "Influencer Marketing",30000, 29000,   1000, ""],
        ["2026-04", "Logistics & Shipping",20000, 22400,  -2400, "Higher COD orders this month"],
        ["2026-04", "Salaries",           45000,  45000,      0, ""],
        ["2026-04", "Rent & Utilities",   18000,  18000,      0, ""],
        ["2026-04", "Miscellaneous",       5000,   3200,   1800, ""],
        ["2026-05", "Raw Materials",      90000,      0,  90000, "Planned — new sandalwood batch"],
        ["2026-05", "Packaging",          35000,      0,  35000, ""],
        ["2026-05", "Meta Ads",           65000,      0,  65000, "Increased for summer campaign"],
        ["2026-05", "Google Ads",         45000,      0,  45000, ""],
        ["2026-05", "Influencer Marketing",35000,     0,  35000, "2 YouTube collabs planned"],
        ["2026-05", "Logistics & Shipping",22000,     0,  22000, ""],
        ["2026-05", "Salaries",           45000,      0,  45000, ""],
        ["2026-05", "Rent & Utilities",   18000,      0,  18000, ""],
        ["2026-05", "Miscellaneous",       5000,      0,   5000, ""],
    ]
)


wb.save("docs/sample_data.xlsx")
print("Created docs/sample_data.xlsx")
print("  Sheet 1: Inventory")
print("  Sheet 2: Raw Material Cost")
print("  Sheet 3: Vendors")
print("  Sheet 4: Monthly Budget")
